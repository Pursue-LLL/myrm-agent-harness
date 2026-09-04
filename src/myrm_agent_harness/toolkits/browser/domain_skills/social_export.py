"""Social data export utility for domain skills.

Exports extracted social media feeds, notes, and comments to Excel (.xlsx) or UTF-8-SIG CSV.

[INPUT]
- records: Sequence[dict[str, Any]] | str (JSON-encoded array of records)
- filename: Optional[str]
- format: "xlsx" | "csv" (default: "xlsx")
- output_dir: Optional[str | Path]

[OUTPUT]
- dict containing status, filepath, row_count, columns, file_size_bytes

[POS]
Domain skill executable layer helper. Pure Python + openpyxl (with CSV fallback).
"""

from __future__ import annotations

import csv
from datetime import datetime
import json
import logging
import os
from pathlib import Path
from typing import Any, Sequence

logger = logging.getLogger(__name__)


def _resolve_export_dir(output_dir: str | Path | None = None) -> Path:
    """Resolve a safe, writable export directory."""
    if output_dir:
        p = Path(output_dir)
        p.mkdir(parents=True, exist_ok=True)
        return p

    if os.path.exists("/workspace"):
        p = Path("/workspace/exports")
    else:
        p = Path.home() / ".myrm" / "exports"

    p.mkdir(parents=True, exist_ok=True)
    return p


def _sanitize_filename(name: str, default_ext: str = "xlsx") -> str:
    """Sanitize filename and ensure valid extension."""
    clean = "".join(c for c in name if c.isalnum() or c in ("-", "_", ".")).strip()
    if not clean:
        clean = f"social_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if not clean.lower().endswith(f".{default_ext}"):
        clean = f"{clean}.{default_ext}"
    return clean


def export_records(
    records: Sequence[dict[str, Any]] | str,
    *,
    filename: str | None = None,
    export_format: str = "xlsx",
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    """Export list of dictionaries to Excel (.xlsx) or CSV.

    Args:
        records: List of dictionaries or JSON string.
        filename: Optional target file name.
        export_format: "xlsx" or "csv".
        output_dir: Optional destination directory.

    Returns:
        Summary dict with filepath, row_count, file_size_bytes, etc.
    """
    if isinstance(records, str):
        try:
            parsed = json.loads(records)
            if isinstance(parsed, list):
                records = parsed
            elif isinstance(parsed, dict):
                records = [parsed]
            else:
                records = []
        except Exception:
            records = []

    normalized_records: list[dict[str, Any]] = [r for r in records if isinstance(r, dict)]
    export_dir = _resolve_export_dir(output_dir)
    fmt = export_format.lower().strip()
    if fmt not in ("xlsx", "csv"):
        fmt = "xlsx"

    # Collect unified fieldnames preserving insertion order
    fieldnames: list[str] = []
    seen_fields: set[str] = set()
    for row in normalized_records:
        for k in row:
            if k not in seen_fields:
                seen_fields.add(k)
                fieldnames.append(k)

    if not fieldnames:
        fieldnames = ["message"]
        if not normalized_records:
            normalized_records = [{"message": "No records extracted"}]

    target_name = _sanitize_filename(filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}", fmt)
    target_path = export_dir / target_name

    # Try Excel export first if requested
    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Social Data"

            # Header row
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="336699", end_color="336699", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.append(fieldnames)
            for col_idx in range(1, len(fieldnames) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Data rows
            for row in normalized_records:
                row_vals = []
                for k in fieldnames:
                    val = row.get(k, "")
                    if isinstance(val, (dict, list)):
                        row_vals.append(json.dumps(val, ensure_ascii=False))
                    else:
                        row_vals.append(str(val) if val is not None else "")
                ws.append(row_vals)

            # Auto-adjust column width
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    # Treat CJK characters as length 2 for width calculation
                    cjk_count = sum(1 for ch in val_str if ord(ch) > 127)
                    cell_len = len(val_str) + cjk_count
                    if cell_len > max_len:
                        max_len = cell_len
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)

            wb.save(target_path)
            return {
                "status": "success",
                "filepath": str(target_path.resolve()),
                "filename": target_name,
                "format": "xlsx",
                "row_count": len(normalized_records),
                "columns": fieldnames,
                "file_size_bytes": target_path.stat().st_size,
            }
        except ImportError:
            logger.warning("openpyxl not available, falling back to UTF-8-SIG CSV export.")
            fmt = "csv"
            target_name = _sanitize_filename(target_name.rsplit(".", 1)[0], "csv")
            target_path = export_dir / target_name

    # CSV export with UTF-8 BOM
    with open(target_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in normalized_records:
            clean_row = {}
            for k in fieldnames:
                val = row.get(k, "")
                if isinstance(val, (dict, list)):
                    clean_row[k] = json.dumps(val, ensure_ascii=False)
                else:
                    clean_row[k] = str(val) if val is not None else ""
            writer.writerow(clean_row)

    return {
        "status": "success",
        "filepath": str(target_path.resolve()),
        "filename": target_name,
        "format": "csv",
        "row_count": len(normalized_records),
        "columns": fieldnames,
        "file_size_bytes": target_path.stat().st_size,
    }
