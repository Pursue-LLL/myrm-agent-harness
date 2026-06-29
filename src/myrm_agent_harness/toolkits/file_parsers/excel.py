"""Excel file parser

Uses openpyxl for parsing XLSX files with support for:
- Merged cells handling
- Markdown/text content output
- Structure metadata output (lightweight overview for large files)
- Audit mode (formula error detection)
- Multiple worksheets

[INPUT]
- (none)

[OUTPUT]
- ExcelParser: Excel file parser using openpyxl

[POS]
Excel file parser
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
from pathlib import Path
from typing import TYPE_CHECKING, Literal

from myrm_agent_harness.toolkits.file_parsers.base import FileParser

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

OutputFormat = Literal["markdown", "text", "structure", "audit"]

_CROSS_SHEET_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_]\w*))!")


class ExcelParser(FileParser):
    """Excel file parser using openpyxl

    Features:
    - Merged cells handling (auto-flatten)
    - Markdown/text table output format
    - Structure mode: lightweight JSON metadata (~500 tokens vs 100k for full dump)
    - Audit mode: deterministic formula error detection
    - Dynamic data_only based on mode (content=True, structure/audit=False)
    """

    def __init__(self, output_format: OutputFormat = "markdown"):
        self._output_format: OutputFormat = output_format

    async def parse(self, file_path: str) -> str:
        """Parse Excel file in the configured output format"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        content = await asyncio.to_thread(self._parse_sync, file_path)

        logger.debug("Excel file parsed: %s, format=%s, length=%d chars", path.name, self._output_format, len(content))
        return content

    def _parse_sync(self, file_path: str) -> str:
        """Synchronously parse Excel file"""
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError("openpyxl is not installed. Run: uv add openpyxl") from e

        use_data_only = self._output_format in ("markdown", "text")
        wb = load_workbook(file_path, data_only=use_data_only)

        if self._output_format == "structure":
            return self._build_structure(wb)
        if self._output_format == "audit":
            return self._build_audit(wb)

        all_text: list[str] = []
        for sheet_name in wb.sheetnames:
            sheet: Worksheet = wb[sheet_name]
            data = self._build_data_matrix(sheet)
            if not data:
                continue

            if self._output_format == "markdown":
                sheet_text = self._format_markdown_table(sheet_name, data)
            else:
                sheet_text = self._format_text_table(sheet_name, data)

            all_text.append(sheet_text)

        return "\n\n".join(all_text)

    # ======================== Structure Mode ========================

    def _build_structure(self, wb: Workbook) -> str:
        """Return lightweight JSON structural metadata for the workbook."""
        sheets_meta: list[dict[str, object]] = []
        cross_sheet_refs: dict[str, list[str]] = {}

        for sheet_name in wb.sheetnames:
            sheet: Worksheet = wb[sheet_name]
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0

            headers: list[str] = []
            if max_row > 0:
                for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=False)):
                    val = cell.value
                    headers.append(str(val) if val is not None else "")

            formula_cols: set[str] = set()
            formula_count = 0
            refs: set[str] = set()
            type_counts: dict[str, int] = {"text": 0, "number": 0, "date": 0, "formula": 0, "empty": 0}

            for row in sheet.iter_rows(min_row=1, values_only=False):
                for cell in row:
                    cell_obj: Cell = cell
                    val = cell_obj.value
                    if val is None:
                        type_counts["empty"] += 1
                    elif isinstance(val, str) and val.startswith("="):
                        type_counts["formula"] += 1
                        formula_count += 1
                        col_letter = cell_obj.column_letter
                        formula_cols.add(col_letter)
                        for match in _CROSS_SHEET_RE.finditer(val):
                            ref_sheet = match.group(1) or match.group(2)
                            if ref_sheet != sheet_name:
                                refs.add(ref_sheet)
                    elif isinstance(val, (int, float)):
                        type_counts["number"] += 1
                    elif hasattr(val, "strftime"):
                        type_counts["date"] += 1
                    else:
                        type_counts["text"] += 1

            merged_regions = [str(r) for r in sheet.merged_cells.ranges]

            sheet_info: dict[str, object] = {
                "name": sheet_name,
                "rows": max_row,
                "cols": max_col,
                "headers": headers[:20],
            }
            if formula_cols:
                sheet_info["formula_columns"] = sorted(formula_cols)
                sheet_info["formula_count"] = formula_count
            if refs:
                sheet_info["cross_references"] = sorted(refs)
                cross_sheet_refs[sheet_name] = sorted(refs)
            if merged_regions:
                sheet_info["merged_regions"] = merged_regions[:10]
            filtered_types = {k: v for k, v in type_counts.items() if v > 0}
            if filtered_types:
                sheet_info["data_types"] = filtered_types

            sheets_meta.append(sheet_info)

        result: dict[str, object] = {"sheets": sheets_meta}
        if cross_sheet_refs:
            result["cross_sheet_references"] = cross_sheet_refs

        return json.dumps(result, ensure_ascii=False, separators=(",", ":"))

    # ======================== Audit Mode ========================

    def _build_audit(self, wb: Workbook) -> str:
        """Return JSON audit findings for formula errors and data issues."""
        findings: list[dict[str, str]] = []

        for sheet_name in wb.sheetnames:
            sheet: Worksheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    cell_obj: Cell = cell
                    val = cell_obj.value
                    if val is None:
                        continue

                    coord = cell_obj.coordinate

                    if isinstance(val, str) and val.startswith("="):
                        self._audit_formula(val, coord, sheet_name, wb.sheetnames, findings)
                    elif isinstance(val, str) and val.startswith("#"):
                        findings.append({
                            "severity": "error",
                            "category": "formula_error",
                            "cell": coord,
                            "sheet": sheet_name,
                            "description": f"Error value: {val}",
                        })

        error_count = sum(1 for f in findings if f["severity"] == "error")
        warning_count = sum(1 for f in findings if f["severity"] == "warning")

        result = {
            "findings": findings[:100],
            "summary": {"errors": error_count, "warnings": warning_count, "total": len(findings)},
        }
        return json.dumps(result, ensure_ascii=False, separators=(",", ":"))

    @staticmethod
    def _audit_formula(
        formula: str,
        coord: str,
        sheet_name: str,
        all_sheets: list[str],
        findings: list[dict[str, str]],
    ) -> None:
        """Check a single formula for common issues."""
        for match in _CROSS_SHEET_RE.finditer(formula):
            ref_sheet = match.group(1) or match.group(2)
            if ref_sheet not in all_sheets:
                findings.append({
                    "severity": "error",
                    "category": "broken_reference",
                    "cell": coord,
                    "sheet": sheet_name,
                    "description": f"References non-existent sheet '{ref_sheet}'",
                    "evidence": formula,
                })

    # ======================== Content Modes ========================

    def _build_data_matrix(self, sheet: Worksheet) -> list[list[str]]:
        """Build data matrix with merged cells handling"""
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        if max_row == 0 or max_col == 0:
            return []

        data: list[list[str]] = [["" for _ in range(max_col)] for _ in range(max_row)]

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            for col_idx, cell in enumerate(row):
                if cell is not None:
                    data[row_idx][col_idx] = str(cell)

        for merged_range in sheet.merged_cells.ranges:
            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            value = str(top_left_cell.value) if top_left_cell.value is not None else ""

            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    data[row - 1][col - 1] = value

        return [row for row in data if any(cell.strip() for cell in row)]

    def _format_markdown_table(self, sheet_name: str, data: list[list[str]]) -> str:
        """Format as Markdown table"""
        if not data:
            return f"## Sheet: {sheet_name}\n\n*Empty*"

        lines = [f"## Sheet: {sheet_name}", ""]

        headers = [self._escape_markdown_cell(cell) for cell in data[0]]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        for row in data[1:]:
            cells = [self._escape_markdown_cell(cell) for cell in row]
            while len(cells) < len(headers):
                cells.append("")
            lines.append("| " + " | ".join(cells) + " |")

        return "\n".join(lines)

    def _format_text_table(self, sheet_name: str, data: list[list[str]]) -> str:
        """Format as plain text (pipe-separated)"""
        lines = [f"## Sheet: {sheet_name}"]
        for row in data:
            lines.append(" | ".join(row))
        return "\n".join(lines)

    @staticmethod
    def _escape_markdown_cell(text: str) -> str:
        """Escape Markdown table cell"""
        if not text:
            return ""
        return text.replace("|", "\\|").replace("\n", " ").replace("\r", " ").strip()

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]
