"""Xlsx formula recalculation check via LibreOffice.

[INPUT]
- Path to .xlsx workbooks touched by bash execution

[OUTPUT]
- run_xlsx_recalc_check: user-facing warnings when recalc finds Excel error cells

[POS]
Optional post-bash formula error detection; complements compare_xlsx_formulas in office_opc.
"""

from __future__ import annotations

import asyncio
import logging
import os
import shutil
import subprocess
import tempfile
import time
from pathlib import Path

logger = logging.getLogger(__name__)

_MACRO_FILENAME = "Module1.xba"
_RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

_EXCEL_ERROR_TOKENS = (
    "#REF!",
    "#NAME?",
    "#DIV/0!",
    "#VALUE!",
    "#NULL!",
    "#NUM!",
    "#N/A",
)
_MAX_ERROR_SAMPLES = 5


def _file_stamp(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_mtime_ns, stat.st_size


def _setup_libreoffice_macro(profile_dir: Path, timeout: int) -> tuple[str | None, str | None]:
    profile_url = profile_dir.as_uri()
    try:
        subprocess.run(
            [
                "soffice",
                "--headless",
                "--terminate_after_init",
                f"-env:UserInstallation={profile_url}",
            ],
            capture_output=True,
            timeout=timeout,
            check=False,
        )
    except FileNotFoundError:
        return None, "soffice not found"
    except subprocess.TimeoutExpired:
        return None, "LibreOffice timed out while preparing recalc profile"

    macro_dir = profile_dir / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        return None, "LibreOffice did not create a usable recalc profile"

    try:
        (macro_dir / _MACRO_FILENAME).write_text(_RECALCULATE_MACRO, encoding="utf-8")
    except OSError as exc:
        return None, "Could not install recalculation macro"

    return profile_url, None


def _collect_excel_error_warnings(file_path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl is not installed; skipping xlsx recalc error scan for %s", file_path)
        return []

    warnings: list[str] = []
    error_locations: list[str] = []

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    if any(token in value for token in _EXCEL_ERROR_TOKENS):
                        error_locations.append(f"{sheet.title}!{cell.coordinate}")
    finally:
        workbook.close()

    if not error_locations:
        return warnings

    sample = ", ".join(error_locations[:_MAX_ERROR_SAMPLES])
    extra = f" (+{len(error_locations) - _MAX_ERROR_SAMPLES} more)" if len(error_locations) > _MAX_ERROR_SAMPLES else ""
    warnings.append(
        f"Excel formula errors detected after recalc ({len(error_locations)} cells, e.g. {sample}{extra}). "
        "Fix broken references before delivery."
    )
    return warnings


def _recalc_workbook_sync(file_path: Path, *, timeout_seconds: int) -> list[str]:
    if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        return []
    if not file_path.is_file():
        return []
    if shutil.which("soffice") is None:
        logger.debug("Xlsx recalc skipped: soffice not available for %s", file_path)
        return []

    if not os.access(file_path, os.W_OK):
        return [f"Excel recalc skipped for {file_path.name}: file is not writable."]

    abs_path = str(file_path.resolve())
    started = time.monotonic()

    with tempfile.TemporaryDirectory(prefix="office-xlsx-recalc-") as profile_root:
        profile_dir = Path(profile_root)
        profile_url, setup_error = _setup_libreoffice_macro(profile_dir, timeout=timeout_seconds)
        if setup_error is not None:
            return [f"Excel recalc failed for {file_path.name}: {setup_error}"]

        remaining_timeout = max(5, timeout_seconds - int(time.monotonic() - started))
        before_stamp = _file_stamp(file_path)
        command = [
            "soffice",
            "--headless",
            "--norestore",
            f"-env:UserInstallation={profile_url}",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]

        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=remaining_timeout + 15,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return [
                f"Excel recalc timed out for {file_path.name} after {remaining_timeout}s; "
                "formula errors may remain undetected."
            ]
        except OSError as exc:
            return [f"Excel recalc failed to run soffice for {file_path.name}"]

        if result.returncode != 0:
            detail = (result.stderr or "").strip() or f"exit {result.returncode}"
            return [f"Excel recalc failed for {file_path.name}: {detail[:200]}"]

        if _file_stamp(file_path) == before_stamp:
            return [
                f"Excel recalc did not rewrite {file_path.name}; formula errors may remain undetected."
            ]

    return _collect_excel_error_warnings(file_path)


async def run_xlsx_recalc_check(file_path: Path, *, timeout_seconds: int = 30) -> list[str]:
    """Recalculate a workbook with LibreOffice and warn on Excel error cells."""
    return await asyncio.to_thread(
        _recalc_workbook_sync,
        file_path,
        timeout_seconds=timeout_seconds,
    )
