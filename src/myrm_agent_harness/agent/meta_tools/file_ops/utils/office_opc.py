"""Office OPC / workbook metrics for write-fidelity audits.

[INPUT]
- Paths to .docx / .xlsx files on disk

[OUTPUT]
- collect_docx_opc_metrics, collect_xlsx_formulas, office_file_audit_read_error,
  extract_office_paths_from_command

[POS]
Shared metrics for bash post-audit and file_ops guards.
"""

from __future__ import annotations

import logging
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

_OFFICE_PATH_RE = re.compile(
    r"""['"]([^'"]+\.(?:docx|doc|xlsx|xls|xlsm))['"]|"""
    r"""(/[^\s'";]+\.(?:docx|doc|xlsx|xls|xlsm))\b|"""
    r"""(\./[^\s'";]+\.(?:docx|doc|xlsx|xls|xlsm))\b|"""
    r"""([^\s'";/\\]+\.(?:docx|doc|xlsx|xls|xlsm))\b""",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class DocxOpcMetrics:
    """Lightweight OPC metrics for detecting destructive rewrites."""

    run_properties_count: int


@dataclass(frozen=True)
class XlsxFormulaSnapshot:
    """Formula cells keyed by sheet, coordinate, and formula text."""

    formulas: frozenset[tuple[str, str, str]]


def extract_office_paths_from_command(command: str) -> list[str]:
    """Return candidate Office file paths mentioned in bash/python command text."""
    seen: set[str] = set()
    ordered: list[str] = []
    for match in _OFFICE_PATH_RE.finditer(command):
        candidate = next(group for group in match.groups() if group)
        normalized = candidate.strip()
        if normalized not in seen:
            seen.add(normalized)
            ordered.append(normalized)
    return ordered


def resolve_office_path(workspace_root: str, candidate: str) -> Path:
    """Resolve a command path against the executor workspace root."""
    path = Path(candidate)
    if path.is_absolute():
        return path
    return (Path(workspace_root) / candidate).resolve()


def office_file_audit_read_error(file_path: Path) -> str | None:
    """Return a user-facing warning when an Office file cannot be audited."""
    suffix = file_path.suffix.lower()
    if not file_path.is_file():
        return None

    if suffix == ".docx":
        try:
            with zipfile.ZipFile(file_path) as archive:
                archive.namelist()
        except zipfile.BadZipFile:
            return (
                f"Office fidelity could not be verified for {file_path.name}: "
                "the file appears corrupt or is not a valid DOCX package."
            )
        except OSError as exc:
            return (
                f"Office fidelity could not be verified for {file_path.name}: "
                f"could not read the file ({exc})."
            )
        return None

    if suffix not in {".xlsx", ".xlsm"}:
        return None

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        logger.warning("openpyxl is not installed; skipping xlsx audit read check for %s", file_path)
        return None

    try:
        workbook = load_workbook(file_path, data_only=False, read_only=True)
    except (zipfile.BadZipFile, InvalidFileException) as exc:
        return (
            f"Office fidelity could not be verified for {file_path.name}: "
            f"the file appears corrupt or is not a valid XLSX package ({exc})."
        )
    except OSError as exc:
        return (
            f"Office fidelity could not be verified for {file_path.name}: "
            f"could not read the workbook ({exc})."
        )

    workbook.close()
    return None


def collect_docx_opc_metrics(file_path: Path) -> DocxOpcMetrics | None:
    """Count run-property nodes in word/document.xml — drops hint at formatting loss."""
    if file_path.suffix.lower() not in {".docx", ".doc"}:
        return None
    if not file_path.is_file():
        return None
    try:
        with zipfile.ZipFile(file_path) as archive:
            if "word/document.xml" not in archive.namelist():
                return None
            document_xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
    except (OSError, zipfile.BadZipFile) as exc:
        logger.warning("Failed to read DOCX OPC metrics for %s: %s", file_path, exc)
        return None

    run_properties_count = document_xml.count("<w:rPr")
    return DocxOpcMetrics(run_properties_count=run_properties_count)


def collect_xlsx_formulas(file_path: Path) -> XlsxFormulaSnapshot | None:
    """Collect formula cells from a workbook (data_only=False)."""
    if file_path.suffix.lower() not in {".xlsx", ".xls", ".xlsm", ".xls"}:
        return None
    if not file_path.is_file():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl is not installed; skipping xlsx formula audit for %s", file_path)
        return None

    try:
        workbook = load_workbook(file_path, data_only=False, read_only=True)
    except OSError as exc:
        logger.warning("Failed to open workbook for formula audit %s: %s", file_path, exc)
        return None

    formulas: set[tuple[str, str, str]] = set()
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas.add((sheet.title, cell.coordinate, value))
    finally:
        workbook.close()

    return XlsxFormulaSnapshot(formulas=frozenset(formulas))


def compare_docx_opc_metrics(
    before: DocxOpcMetrics | None,
    after: DocxOpcMetrics | None,
) -> list[str]:
    """Return user-facing warnings when formatting-related OPC nodes drop sharply."""
    if before is None or after is None:
        return []
    if before.run_properties_count == 0:
        return []
    if after.run_properties_count >= before.run_properties_count:
        return []
    lost = before.run_properties_count - after.run_properties_count
    if lost < max(3, before.run_properties_count // 4):
        return []
    return [
        (
            "DOCX formatting may have been degraded "
            f"(run properties {before.run_properties_count} → {after.run_properties_count}). "
            "Prefer XML-level <w:t> edits per office-document skill."
        )
    ]


def compare_xlsx_formulas(
    before: XlsxFormulaSnapshot | None,
    after: XlsxFormulaSnapshot | None,
) -> list[str]:
    """Return warnings when formulas disappear after an edit."""
    if before is None or after is None:
        return []
    if not before.formulas:
        return []
    missing = before.formulas - after.formulas
    if not missing:
        return []
    sample = ", ".join(f"{sheet}!{coord}" for sheet, coord, _ in list(missing)[:5])
    extra = f" (+{len(missing) - 5} more)" if len(missing) > 5 else ""
    return [
        (
            f"Excel formulas were removed or overwritten ({len(missing)} cells, e.g. {sample}{extra}). "
            "Use openpyxl cell writes without replacing '=' cells."
        )
    ]
